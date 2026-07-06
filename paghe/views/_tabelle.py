"""Modulo _tabelle - generato automaticamente da views.py"""
from django.urls import reverse
from paghe.views._common_imports import *
import logging
logger = logging.getLogger(__name__)
from paghe.views._helpers import _resolve_valore, _xlsx_field_to_python, _get_tabella_conf, _dati_eliminato, _parse_toggles, _parse_convivenza_items
from paghe.views._constants import XLSX_EXPORT_MAP, XLSX_IMPORT_MAP
from paghe.views._calcoli_core import _calcola_busta_data
from paghe.views._gestione_pdf import _registra_font_pdf
from paghe.models import RecordEliminato, TabellaScattiAnzianita, ParametriCCNL, Livello, ContrattoAttivo

@login_required
@never_cache
def tabella_list(request, tipo):
    conf = _get_tabella_conf(tipo)
    if not conf:
        return render(request, 'paghe/tabella_generica.html', {'error': 'Tabella non trovata'})
    model = conf['model']
    objects = model.objects.all()
    meta_fields = model._meta.get_fields()
    fields = []
    for f in meta_fields:
        if not hasattr(f, 'name') or f.name.startswith('_') or getattr(f, 'auto_created', False) or f.name in ('id', 'pk'):
            continue
        ftype = f.get_internal_type() if hasattr(f, 'get_internal_type') else ''
        if ftype == 'ManyToManyField' and not getattr(f, 'related_name', ''):
            continue
        fields.append({'name': f.name, 'verbose_name': getattr(f, 'verbose_name', f.name) or f.name, 'type': ftype})
    ctx = {
        'opzioni': get_opzioni(),
        'titolo': conf.get('titolo', tipo),
        'icona': conf.get('icona', 'bi-table'),
        'objects': objects,
        'fields': fields,
        'model_name': model.__name__,
        'verbose_name': model._meta.verbose_name_plural if hasattr(model._meta, 'verbose_name_plural') else model.__name__,
        'tipo_url': tipo,
        'colori_livelli': {},
    }
    if tipo in ('parametri-ccnl', 'scatti-anzianita', 'livelli'):
        ctx['colori_livelli'] = {l.codice: l.colore for l in Livello.objects.all()}
    if tipo == 'parametri-ccnl':
        ctx['keep_fields'] = ['livello', 'descrizione_corta', 'descrizione_lunga', 'paga_base', 'anno']
        ctx['truncate_fields'] = ['descrizione_lunga']
        ctx['anni'] = list(ParametriCCNL.objects.values_list('anno', flat=True).distinct().order_by('-anno'))
    if tipo == 'progetti-regionali':
        return render(request, 'paghe/progetti_regionali_list.html', ctx)
    if tipo == 'modelli-lista':
        return render(request, 'paghe/modelli_lista_list.html', ctx)
    return render(request, 'paghe/tabella_generica.html', ctx)

@login_required
@permesso_richiesto('impostazioni')
def ajax_nuovo_tabella(request, tipo):
    conf = _get_tabella_conf(tipo)
    if not conf:
        return JsonResponse({'success': False, 'message': 'Tabella non trovata'})
    model = conf['model']
    form_class = conf['form_class']
    opz = get_opzioni()
    TEMPLATE_MAP = {
        'progetti-regionali': 'frontend/ajax_form_progetto.html',
        'tabelle-casse': 'frontend/ajax_form_tabella_numeri.html',
        'contributi-inps': 'frontend/ajax_form_tabella_numeri.html',
        'parametri-ccnl': 'frontend/ajax_form_parametri_ccnl.html',
    }
    template = TEMPLATE_MAP.get(tipo, 'frontend/ajax_form_tabella.html')
    if request.method == 'POST':
        if any(k.startswith('file_') for k in request.FILES):
            form = form_class(request.POST, request.FILES)
        else:
            form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'redirect_url': reverse('tabella_list', args=[tipo])})
        return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)
    form = form_class()
    ctx = {
        'form': form, 'titolo': conf.get('titolo', tipo),
        'model_name': model.__name__, 'tipo': tipo,
        'tiny_api_key': getattr(opz, 'tiny_api_key', '') if opz else '',
    }
    html = render(request, template, ctx).content.decode('utf-8')
    return JsonResponse({'success': True, 'html': html})

@login_required
@permesso_richiesto('impostazioni')
def ajax_modifica_tabella(request, tipo, pk):
    conf = _get_tabella_conf(tipo)
    if not conf:
        return JsonResponse({'success': False, 'message': 'Tabella non trovata'})
    model = conf['model']
    form_class = conf['form_class']
    obj = get_object_or_404(model, pk=pk)
    opz = get_opzioni()
    TEMPLATE_MAP = {
        'progetti-regionali': 'frontend/ajax_form_progetto.html',
        'tabelle-casse': 'frontend/ajax_form_tabella_numeri.html',
        'contributi-inps': 'frontend/ajax_form_tabella_numeri.html',
        'parametri-ccnl': 'frontend/ajax_form_parametri_ccnl.html',
    }
    template = TEMPLATE_MAP.get(tipo, 'frontend/ajax_form_tabella.html')
    ctx = {
        'instance': obj, 'titolo': conf.get('titolo', tipo),
        'model_name': model.__name__, 'tipo': tipo,
        'tiny_api_key': getattr(opz, 'tiny_api_key', '') if opz else '',
    }
    if tipo == 'progetti-regionali':
        progetti_qs = ContrattoAttivo.objects.filter(progetto=obj).values_list('pk', 'datore__nome_cognome', 'lavoratore__nome_cognome')
        list(progetti_qs)
        progetti_nav_data = [
            {'pk': p.pk, 'label': f"{p.beneficiario.nome_cognome} - {p.tipo.nome if p.tipo else 'N/D'}"}
            for p in model.objects.all() if hasattr(p, 'beneficiario')
        ]
        ctx['progetti_nav_data'] = progetti_nav_data
        ctx['progetto_corrente_pk'] = obj.pk
    if request.method == 'POST':
        if any(k.startswith('file_') for k in request.FILES):
            form = form_class(request.POST, request.FILES, instance=obj)
        else:
            form = form_class(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'redirect_url': reverse('tabella_list', args=[tipo])})
        return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)
    form = form_class(instance=obj)
    ctx['form'] = form
    html = render(request, template, ctx).content.decode('utf-8')
    return JsonResponse({'success': True, 'html': html})

@login_required
@permesso_richiesto('impostazioni')
def ajax_duplica_tabella(request, tipo, pk):
    conf = _get_tabella_conf(tipo)
    if not conf:
        return JsonResponse({'success': False, 'message': 'Tabella non trovata'})
    model = conf['model']
    form_class = conf['form_class']
    obj = get_object_or_404(model, pk=pk)
    opz = get_opzioni()
    template = 'frontend/ajax_form_tabella.html'
    initial = {}
    for f in model._meta.get_fields():
        if hasattr(f, 'name') and not getattr(f, 'auto_created', False) and not getattr(f, 'primary_key', False):
            initial[f.name] = getattr(obj, f.name, None)
    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'redirect_url': reverse('tabella_list', args=[tipo])})
        return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)
    form = form_class(initial=initial)
    ctx = {
        'form': form, 'titolo': f"Duplica {conf.get('titolo', tipo)}",
        'model_name': model.__name__, 'tipo': tipo,
        'tiny_api_key': getattr(opz, 'tiny_api_key', '') if opz else '',
    }
    html = render(request, template, ctx).content.decode('utf-8')
    return JsonResponse({'success': True, 'html': html})

@login_required
@permesso_richiesto('impostazioni')
@require_http_methods(['POST'])
def ajax_elimina_tabella(request, tipo, pk):
    conf = _get_tabella_conf(tipo)
    if not conf:
        return JsonResponse({'success': False, 'message': 'Tabella non trovata'})
    model = conf['model']
    obj = get_object_or_404(model, pk=pk)
    RecordEliminato.objects.create(
        modello=f"{model.__name__}({pk})",
        oggetto_id=str(pk),
        dati_json=_dati_eliminato(obj),
    )
    obj.delete()
    return JsonResponse({'success': True})

@login_required
def ajax_esporta_xlsx(request, tipo):
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    mapping = XLSX_EXPORT_MAP.get(tipo)
    if not mapping:
        return JsonResponse({'success': False, 'message': 'Export non configurato per questo tipo'})
    model_class = mapping['model']
    fields = mapping['fields']
    qs = model_class.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo
    header_fill = PatternFill(start_color='5E6AD2', end_color='5E6AD2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx, field_name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=mapping.get('headers', {}).get(field_name, field_name))
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, obj in enumerate(qs, 2):
        for col_idx, field_name in enumerate(fields, 1):
            val = _resolve_valore(obj, field_name)
            if isinstance(val, date):
                val = val.strftime('%d/%m/%Y')
            elif isinstance(val, bool):
                val = 'Si' if val else 'No'
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')
    from datetime import datetime
    oggi_str = datetime.now().strftime('%Y%m%d')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="anagrafe_{tipo}_{oggi_str}.xlsx"'
    wb.save(response)
    return response

@login_required
@require_http_methods(['POST'])
def ajax_importa_xlsx(request, tipo):
    import openpyxl
    mapping = XLSX_IMPORT_MAP.get(tipo)
    if not mapping:
        return JsonResponse({'success': False, 'message': 'Import non configurato per questo tipo'})
    model_class = mapping['model']
    form_class = mapping.get('form_class')
    field_map = mapping.get('fields', {})
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'Nessun file caricato'})
    wb = openpyxl.load_workbook(request.FILES['file'])
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return JsonResponse({'success': False, 'message': 'File vuoto'})
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    importati = 0
    errori = 0
    messages = []
    for row in rows[1:]:
        data = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row) and header in field_map:
                val = _xlsx_field_to_python(row[col_idx])
                data[field_map[header]] = val
        if tipo == 'contratti':
            cf_datore = data.pop('datore__codice_fiscale', None) or data.pop('datore__nome_cognome', None)
            cf_lavoratore = data.pop('lavoratore__codice_fiscale', None) or data.pop('lavoratore__nome_cognome', None)
            if cf_datore and cf_lavoratore:
                from paghe.models import DatoreLavoro, Lavoratore
                datore = DatoreLavoro.objects.filter(codice_fiscale=cf_datore).first()
                if not datore:
                    datore = DatoreLavoro.objects.filter(nome_cognome=cf_datore).first()
                lavoratore = Lavoratore.objects.filter(codice_fiscale=cf_lavoratore).first()
                if not lavoratore:
                    lavoratore = Lavoratore.objects.filter(nome_cognome=cf_lavoratore).first()
                if datore and lavoratore:
                    data['datore'] = datore.pk
                    data['lavoratore'] = lavoratore.pk
        elif tipo == 'progetti-regionali':
            cf_beneficiario = data.pop('beneficiario__codice_fiscale', None) or data.pop('beneficiario__nome_cognome', None)
            tipo_nome = data.pop('tipo__nome', None)
            if cf_beneficiario:
                from paghe.models import Beneficiario, TipoProgettoRegionale
                ben = Beneficiario.objects.filter(codice_fiscale=cf_beneficiario).first()
                if not ben:
                    ben = Beneficiario.objects.filter(nome_cognome=cf_beneficiario).first()
                if ben:
                    data['beneficiario'] = ben.pk
                if tipo_nome:
                    tipi = TipoProgettoRegionale.objects.filter(nome=tipo_nome).first()
                    if tipi:
                        data['tipo'] = tipi.pk
        try:
            if form_class:
                form = form_class(data)
                if form.is_valid():
                    form.save()
                    importati += 1
                else:
                    errori += 1
                    messages.append(f"Errore validazione: {form.errors}")
            else:
                model_class.objects.create(**data)
                importati += 1
        except Exception as e:
            errori += 1
            messages.append(str(e))
    return JsonResponse({'success': True, 'message': f'Importati: {importati}, Errori: {errori}', 'importati': importati, 'errori': errori})

# === CCNL OCCHIO ===

@login_required
def ajax_ccnl_occhio(request):
    parametri_pk = request.GET.get('parametri_pk')
    contratto_pk = request.GET.get('contratto_pk')
    livello_pk = request.GET.get('livello_pk')
    anno_raw = request.GET.get('anno', '').strip()
    if not anno_raw:
        return JsonResponse({'error': 'Anno obbligatorio'}, status=400)
    try:
        anno = int(anno_raw)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Anno non valido'}, status=400)
    contratto = None
    if contratto_pk:
        contratto = get_object_or_404(ContrattoAttivo, pk=contratto_pk)
    parametri = None
    if parametri_pk:
        parametri = get_object_or_404(ParametriCCNL, pk=parametri_pk)
    elif contratto and hasattr(contratto, 'parametri_minimi') and contratto.parametri_minimi:
        parametri = contratto.parametri_minimi
    elif livello_pk:
        livello = get_object_or_404(Livello, pk=livello_pk)
        parametri = ParametriCCNL.objects.filter(livello=livello, anno=anno).first()
    if not parametri:
        return JsonResponse({'error': 'Parametri CCNL non trovati'}, status=404)

    def _attiva(field_name):
        return getattr(contratto, field_name, False) if contratto else None

    def _ind_value(value, attiva):
        return float(value) if attiva and value else 0.0

    paga_base = float(parametri.paga_base) if parametri.paga_base else 0.0
    indennita_items = []
    ind_map = [
        ('ind_certificazione_qualita', 'Certificazione Qualità', 'ind_cert_qualita', 'ind_certificazione_qualita'),
        ('ind_piu_persone_non_conv', 'Assistenza Più Persone', 'ind_assistenza_piu_persone_non_conv', 'ind_piu_persone_non_conv'),
        ('ind_minori_non_conv', 'Minori < 6 Anni', 'ind_minori_6_anni_non_conv', 'ind_minori_non_conv'),
        ('ind_piu_persone_qualita', 'Più Persone + Qualità', 'ind_piu_persone_qualita', 'ind_piu_persone_qualita'),
        ('ind_minori_qualita', 'Minori + Qualità', 'ind_minori_qualita', 'ind_minori_qualita'),
        ('ind_funzione_conviventi', 'Funzione Conviventi', 'ind_funzione_conviventi', 'ind_funzione_conviventi'),
        ('ind_assistenza_piu_persone_ft', 'Più Persone (FT)', 'ind_assistenza_piu_persone_ft', 'ind_assistenza_piu_persone_ft'),
        ('ind_assistenza_piu_persone_pt', 'Più Persone (PT)', 'ind_assistenza_piu_persone_pt', 'ind_assistenza_piu_persone_pt'),
        ('ind_minori_6_anni_ft', 'Minori < 6 Anni (FT)', 'ind_minori_6_anni_ft', 'ind_minori_6_anni_ft'),
    ]
    for key, label, param_field, attiva_field in ind_map:
        val = getattr(parametri, param_field, 0) or 0
        attiva = _attiva(attiva_field)
        indennita_items.append({'key': key, 'label': label, 'value': float(val), 'attiva': attiva is True})

    bienni = []
    scatto = TabellaScattiAnzianita.objects.filter(
        livello=parametri.livello.codice if parametri.livello else ''
    ).first()
    scatto_value = float(scatto.valore_scatto) if scatto else 0
    for i in range(7):
        bienni.append({'biennio': i + 1, 'anni': (i + 1) * 2, 'orario': scatto_value * (i + 1), 'mensile': scatto_value * (i + 1) * 170})

    return JsonResponse({
        'paga_base': paga_base,
        'parametri_pk': parametri.pk,
        'indennita': indennita_items,
        'bienni': bienni,
        'scatto_value': scatto_value,
    })


def _costruisci_voci_ccnl(contratto, parametri):
    def _attiva(field_name):
        return getattr(contratto, field_name, False) if contratto else None

    def _valore(param_field):
        return float(getattr(parametri, param_field, 0) or 0)

    paga_base = _valore('paga_base')
    retrib_sost_value = _valore('retribuzione_sostituzione')
    retrib_sost_attiva = _attiva('usa_retribuzione_sostituzione')
    paga_base_effettiva = retrib_sost_value if (retrib_sost_attiva and retrib_sost_value > 0) else paga_base

    retrib_sost_item = {'key': 'retrib_sostituzione', 'label': 'Retrib. Sostituzione', 'value': retrib_sost_value, 'attiva': retrib_sost_attiva}

    ratei = [
        {'key': 'rateo_13ma', 'label': 'Tredicesima Oraria', 'value': _valore('tredicesima_oraria'), 'attiva': _attiva('paga_13ma')},
        {'key': 'rateo_ferie', 'label': 'Ferie Orarie', 'value': _valore('ferie_orarie'), 'attiva': _attiva('paga_ferie')},
        {'key': 'rateo_festivi', 'label': 'Festivi Orari', 'value': _valore('festivi_orari'), 'attiva': _attiva('paga_festivi')},
    ]

    tfr_value = _valore('tfr_orario')
    if contratto:
        tfr_attiva = contratto.modalita_tfr in ('INCLUSO', 'SEPARATO_70')
        tfr_separato_attiva = contratto.modalita_tfr != 'INCLUSO'
        anticipo_70_attiva = contratto.modalita_tfr == 'SEPARATO_70'
    else:
        tfr_attiva = None
        tfr_separato_attiva = False
        anticipo_70_attiva = False

    ratei.append({'key': 'rateo_tfr', 'label': 'Tfr Orario', 'value': tfr_value, 'attiva': tfr_attiva})
    ratei.append({'key': 'rateo_tfr_separato', 'label': 'TFR Separato', 'value': 0, 'attiva': tfr_separato_attiva})
    ratei.append({'key': 'rateo_anticipo_70', 'label': 'Anticipo 70%', 'value': 0, 'attiva': anticipo_70_attiva})

    ind_mapping = [
        ('ind_certificazione_qualita', 'Certificazione Qualità', 'ind_cert_qualita'),
        ('ind_piu_persone_non_conv', 'Assistenza Più Persone', 'ind_assistenza_piu_persone_non_conv'),
        ('ind_minori_non_conv', 'Minori < 6 Anni', 'ind_minori_6_anni_non_conv'),
        ('ind_piu_persone_qualita', 'Più Persone + Qualità', 'ind_piu_persone_qualita'),
        ('ind_minori_qualita', 'Minori + Qualità', 'ind_minori_qualita'),
        ('ind_funzione_conviventi', 'Funzione Conviventi', 'ind_funzione_conviventi'),
        ('ind_assistenza_piu_persone_ft', 'Più Persone (FT)', 'ind_assistenza_piu_persone_ft'),
        ('ind_assistenza_piu_persone_pt', 'Più Persone (PT)', 'ind_assistenza_piu_persone_pt'),
        ('ind_minori_6_anni_ft', 'Minori < 6 Anni (FT)', 'ind_minori_6_anni_ft'),
    ]
    indennita = [{'key': k, 'label': l, 'value': _valore(v), 'attiva': _attiva(k)} for k, l, v in ind_mapping]

    scatto = TabellaScattiAnzianita.objects.filter(livello=parametri.livello.codice if parametri.livello else '').first()
    scatto_value = float(scatto.valore_scatto) if scatto else 0
    scatti_attiva = _attiva('applica_scatti')

    bienni_accumulati = 0
    if contratto and contratto.data_assunzione:
        oggi = date.today()
        anni = oggi.year - contratto.data_assunzione.year
        if (oggi.month, oggi.day) < (contratto.data_assunzione.month, contratto.data_assunzione.day):
            anni -= 1
        bienni_accumulati = min(anni // 2, 7)

    scatto_effettivo = scatto_value * bienni_accumulati if scatti_attiva else 0

    groups = [
        {'title': 'Retribuzione Base', 'icon': 'currency-euro', 'items': [
            {'label': 'Paga Base Oraria', 'value': paga_base, 'fixed': True},
            retrib_sost_item,
        ]},
        {'title': 'Indennità', 'icon': 'stars', 'items': indennita},
        {'title': 'Ratei / Mensilità', 'icon': 'calendar-check', 'items': ratei},
    ]
    all_toggleable = {}
    for item in ratei + indennita + [retrib_sost_item]:
        all_toggleable[item['key']] = item

    return (groups, all_toggleable, scatto_value, scatto_effettivo, scatti_attiva,
            paga_base, paga_base_effettiva, retrib_sost_value, retrib_sost_attiva, bienni_accumulati)


_IND_MAPPING_ORE = [
    ('ind_certificazione_qualita', 'ind_cert_qualita'),
    ('ind_piu_persone_non_conv', 'ind_assistenza_piu_persone_non_conv'),
    ('ind_minori_non_conv', 'ind_minori_6_anni_non_conv'),
    ('ind_piu_persone_qualita', 'ind_piu_persone_qualita'),
    ('ind_minori_qualita', 'ind_minori_qualita'),
    ('ind_assistenza_piu_persone_ft', 'ind_assistenza_piu_persone_ft'),
    ('ind_assistenza_piu_persone_pt', 'ind_assistenza_piu_persone_pt'),
    ('ind_minori_6_anni_ft', 'ind_minori_6_anni_ft'),
    ('ind_funzione_conviventi', 'ind_funzione_conviventi'),
]

_ORE_TOGGLE_MAP = {
    'rateo_13ma': 'paga_13ma',
    'rateo_ferie': 'paga_ferie',
    'rateo_festivi': 'paga_festivi',
    'scatti': 'applica_scatti',
}


def _calcola_ore_simulate(contratto, parametri, toggles_dict, scatto_effettivo):
    """Calcola ore = ceil(budget / paga_oraria_eff).
    toggles_dict: chiavi = nome campo contratto (paga_13ma, ind_certificazione_qualita, ...),
                  valori = bool. Se chiave assente, usa getattr(contratto, chiave).
    """
    budget = contratto.budget_di_partenza
    po = float(parametri.paga_base)

    if toggles_dict.get('paga_13ma', contratto.paga_13ma):
        po += float(parametri.tredicesima_oraria)
    if toggles_dict.get('paga_ferie', contratto.paga_ferie):
        po += float(parametri.ferie_orarie)
    if toggles_dict.get('paga_festivi', contratto.paga_festivi):
        po += float(parametri.festivi_orari)

    mt = contratto.modalita_tfr
    if 'rateo_tfr' in toggles_dict:
        mt = 'INCLUSO' if toggles_dict['rateo_tfr'] else ('SEPARATO_70' if toggles_dict.get('rateo_anticipo_70') else 'SEPARATO')
    elif 'rateo_anticipo_70' in toggles_dict:
        mt = 'SEPARATO_70' if toggles_dict['rateo_anticipo_70'] else 'SEPARATO'

    if mt == 'INCLUSO':
        po += float(parametri.tfr_orario)
    elif mt == 'SEPARATO_70':
        po += float(parametri.tfr_orario) * 0.7

    for contratto_field, param_field in _IND_MAPPING_ORE:
        if toggles_dict.get(contratto_field, getattr(contratto, contratto_field, False)):
            po += float(getattr(parametri, param_field, 0) or 0)

    if toggles_dict.get('applica_scatti', contratto.applica_scatti) and contratto.data_assunzione:
        po += float(scatto_effettivo)

    return math.ceil(budget / po) if po > 0 and budget > 0 else 0


@login_required
def ccnl_occhio_popup(request):
    contratto_pk = request.GET.get('contratto_pk')
    livello_pk = request.GET.get('livello_pk')
    anno_raw = request.GET.get('anno', '').strip()
    try:
        anno = int(anno_raw) if anno_raw else date.today().year
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Anno non valido'}, status=400)

    contratto = None
    if contratto_pk:
        contratto = get_object_or_404(ContrattoAttivo, pk=contratto_pk)

    if livello_pk:
        livello = get_object_or_404(Livello, pk=livello_pk)
        parametri = ParametriCCNL.objects.filter(livello=livello, anno=anno).first()
    elif contratto and hasattr(contratto, 'parametri_minimi') and contratto.parametri_minimi:
        parametri = contratto.parametri_minimi
    else:
        return JsonResponse({'error': 'Specificare contratto_pk o livello_pk + anno'}, status=400)

    if not parametri:
        return JsonResponse({'error': 'Parametri CCNL non trovati'}, status=404)

    groups, all_toggleable, scatto_value, scatto_effettivo, scatti_attiva, paga_base, paga_base_effettiva, retrib_sost_value, retrib_sost_attiva, bienni_accumulati = _costruisci_voci_ccnl(contratto, parametri)

    livelli = Livello.objects.all().order_by('codice')
    livello_contratto = parametri.livello

    mese_corrente = request.GET.get('mese', date.today().month)
    try:
        mese_corrente = int(mese_corrente)
    except (ValueError, TypeError):
        mese_corrente = date.today().month
    ore_simulate = None
    if contratto:
        ore_toggles = {}
        ore_simulate = _calcola_ore_simulate(contratto, parametri, ore_toggles, scatto_effettivo)
    busta_data = _calcola_busta_data(contratto, mese_corrente, anno, parametri_override=parametri, ore_override=ore_simulate) if contratto else {}
    if busta_data and not busta_data.get('errore') and contratto and ore_simulate is not None:
        ore_reali = float(contratto.ore_mensili_calcolate)
        po = busta_data.get('paga_oraria_lorda', float(parametri.paga_base))
        busta_data['differenza'] = round((ore_simulate - ore_reali) * po, 2)

    totale_toggleabili = sum(
        item['value'] for item in all_toggleable.values()
        if item.get('attiva') and item.get('key') != 'retrib_sostituzione'
    ) if all_toggleable else 0

    paga_oraria_lorda = busta_data.get('paga_oraria_lorda', 0) if busta_data else 0

    conv_pranzo_rate = float(parametri.convivenza_pranzo) if parametri else 0
    conv_cena_rate = float(parametri.convivenza_cena) if parametri else 0
    conv_alloggio_rate = float(parametri.convivenza_alloggio) if parametri else 0

    bienni = []
    if contratto and contratto.data_assunzione:
        oggi = date.today()
        anni_diff = oggi.year - contratto.data_assunzione.year
        if (oggi.month, oggi.day) < (contratto.data_assunzione.month, contratto.data_assunzione.day):
            anni_diff -= 1
        max_bienni = min(anni_diff // 2, 7)
        for b in range(1, max_bienni + 1):
            bienni.append({
                'biennio': b,
                'anni': b * 2,
                'orario': round(scatto_value * b, 4),
                'mensile': round(scatto_value * b * 170, 2),
            })

    return render(request, 'paghe/popup_ccnl_occhio.html', {
        'contratto': contratto, 'parametri': parametri, 'livelli': livelli,
        'livello_contratto': livello_contratto, 'groups': groups,
        'all_toggleable': all_toggleable, 'bienni_accumulati': bienni_accumulati,
        'scatto_value': scatto_value, 'scatto_effettivo': scatto_effettivo,
        'scatti_attiva': scatti_attiva, 'paga_base': paga_base,
        'paga_base_effettiva': paga_base_effettiva, 'anno': anno,
        'mese_corrente': mese_corrente, 'busta_json': busta_data,
        'retrib_sost_value': retrib_sost_value, 'retrib_sost_attiva': retrib_sost_attiva,
        'totale_toggleabili': totale_toggleabili, 'paga_oraria_lorda': paga_oraria_lorda,
        'conv_pranzo_rate': conv_pranzo_rate, 'conv_cena_rate': conv_cena_rate,
        'conv_alloggio_rate': conv_alloggio_rate, 'bienni': bienni,
    })


@login_required
def ajax_ccnl_occhio_busta(request, pk):
    contratto = get_object_or_404(ContrattoAttivo, pk=pk)
    livello_pk = request.GET.get('livello_pk')
    mese = int(request.GET.get('mese', date.today().month))
    anno = int(request.GET.get('anno', date.today().year))

    parametri = getattr(contratto, 'parametri_minimi', None)
    if livello_pk:
        liv = get_object_or_404(Livello, pk=livello_pk)
        liv_parametri = ParametriCCNL.objects.filter(livello=liv, anno=anno).first()
        if liv_parametri:
            parametri = liv_parametri
    if not parametri:
        return JsonResponse({'errore': 'Parametri CCNL non associati'})
    groups, all_toggleable, scatto_value, scatto_effettivo, scatti_attiva, paga_base, paga_base_effettiva, retrib_sost_value, retrib_sost_attiva, bienni_accumulati = _costruisci_voci_ccnl(contratto, parametri)

    toggles_dict = _parse_toggles(request) or {}
    convivenza_items = _parse_convivenza_items(request)

    CCNL_TO_INTERNAL = {
        'ind_funzione_conviventi': 'ind_funzione',
        'ind_minori_6_anni_ft': 'ind_bambini_6',
        'ind_assistenza_piu_persone_ft': 'ind_piu_assistiti',
        'ind_certificazione_qualita': 'ind_cert_qualita',
    }
    for ccnl_key, internal_key in CCNL_TO_INTERNAL.items():
        v = request.GET.get(ccnl_key)
        if v is not None:
            toggles_dict[internal_key] = v == '1'

    EXTRA_IND = [
        ('ind_piu_persone_non_conv', 'ind_assistenza_piu_persone_non_conv'),
        ('ind_minori_non_conv', 'ind_minori_6_anni_non_conv'),
        ('ind_piu_persone_qualita', 'ind_piu_persone_qualita'),
        ('ind_minori_qualita', 'ind_minori_qualita'),
        ('ind_assistenza_piu_persone_pt', 'ind_assistenza_piu_persone_pt'),
    ]

    ore_toggles = {}
    for popup_key, model_key in _ORE_TOGGLE_MAP.items():
        v = request.GET.get(popup_key)
        if v is not None:
            ore_toggles[model_key] = v == '1'
    v_tfr = request.GET.get('rateo_tfr')
    v_ant70 = request.GET.get('rateo_anticipo_70')
    if v_tfr is not None:
        ore_toggles['rateo_tfr'] = v_tfr == '1'
    if v_ant70 is not None:
        ore_toggles['rateo_anticipo_70'] = v_ant70 == '1'
    for cf, _ in _IND_MAPPING_ORE:
        v = request.GET.get(cf)
        if v is not None:
            ore_toggles[cf] = v == '1'
    ore_simulate = _calcola_ore_simulate(contratto, parametri, ore_toggles, scatto_effettivo)
    busta_data = _calcola_busta_data(
        contratto, mese=mese, anno=anno,
        toggles=toggles_dict, convivenza_items=convivenza_items,
        parametri_override=parametri, ore_override=ore_simulate,
    )

    extra_indennita = []
    extra_totale = 0.0
    for toggle_key, param_field in EXTRA_IND:
        attiva = request.GET.get(toggle_key)
        if attiva is None:
            attiva = getattr(contratto, toggle_key, False)
        else:
            attiva = attiva == '1'
        if attiva:
            v = float(getattr(parametri, param_field, 0) or 0)
            extra_indennita.append({'label': toggle_key, 'importo': v})
            extra_totale += v

    if extra_indennita:
        busta_data.setdefault('indennita', []).extend(extra_indennita)
        busta_data['totale_indennita'] = round(busta_data.get('totale_indennita', 0) + extra_totale, 4)
        busta_data['totale_lordo'] = round(busta_data.get('totale_lordo', 0) + extra_totale, 2)
        busta_data['netto'] = round(busta_data.get('netto', 0) + extra_totale, 4)
        busta_data['differenza'] = round(busta_data.get('differenza', 0) + extra_totale, 2)

    if busta_data and not busta_data.get('errore'):
        ore_reali = float(contratto.ore_mensili_calcolate)
        po = busta_data.get('paga_oraria_lorda', float(parametri.paga_base))
        busta_data['differenza'] = round((ore_simulate - ore_reali) * po, 2)

    response = {
        'paga_base': paga_base,
        'paga_base_effettiva': paga_base_effettiva,
        'scatto_effettivo': scatto_effettivo,
        'bienni_accumulati': bienni_accumulati,
        'scatti_attiva': scatti_attiva,
    }
    if busta_data:
        response.update(busta_data)
    return JsonResponse(response)


@login_required
def ajax_ccnl_occhio_riepilogo_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.platypus.flowables import HRFlowable as HRFlow

    _registra_font_pdf()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)

    acciaio = HexColor('#2c5282')
    grigio_scuro = HexColor('#222222')
    grigio_bordo = HexColor('#cccccc')
    HexColor('#f0f4f8')

    s_titolo = ParagraphStyle('Titolo', fontName='Roboto-Bold', fontSize=16, textColor=acciaio, spaceAfter=4, alignment=TA_CENTER)
    s_sottotitolo = ParagraphStyle('Sottotitolo', fontName='Roboto', fontSize=10, textColor=HexColor('#555'), spaceAfter=12, alignment=TA_CENTER)
    s_sezione = ParagraphStyle('Sezione', fontName='Roboto-Bold', fontSize=12, textColor=acciaio, spaceBefore=10, spaceAfter=6)
    s_cella = ParagraphStyle('Cella', fontName='Roboto', fontSize=8.5, textColor=grigio_scuro, leading=11)
    s_cella_dx = ParagraphStyle('CellaDx', fontName='Roboto', fontSize=8.5, textColor=grigio_scuro, alignment=TA_RIGHT, leading=11)
    s_cella_c = ParagraphStyle('CellaC', fontName='Roboto', fontSize=8.5, textColor=grigio_scuro, alignment=TA_CENTER, leading=11)
    s_totale = ParagraphStyle('Totale', fontName='Roboto-Bold', fontSize=9, textColor=grigio_scuro, leading=11)
    s_totale_dx = ParagraphStyle('TotaleDx', fontName='Roboto-Bold', fontSize=9, textColor=grigio_scuro, alignment=TA_RIGHT, leading=11)
    ParagraphStyle('SezBusta', fontName='Roboto-Bold', fontSize=9, textColor=grigio_scuro, leading=11)
    s_sez_busta_dx = ParagraphStyle('SezBustaDx', fontName='Roboto-Bold', fontSize=9, textColor=grigio_scuro, alignment=TA_RIGHT, leading=11)
    ParagraphStyle('BustaVal', fontName='Roboto', fontSize=9, textColor=grigio_scuro, leading=11)
    s_busta_val_dx = ParagraphStyle('BustaValDx', fontName='Roboto', fontSize=9, textColor=grigio_scuro, alignment=TA_RIGHT, leading=11)

    def pr(v):
        return Paragraph(str(v) if v is not None else '', s_cella)

    def prd(v):
        return Paragraph(f'€ {v:.4f}' if v else '—', s_cella_dx)

    def prs(attivo):
        return Paragraph('<font color="green">✓</font>' if attivo else '<font color="gray">✗</font>', s_cella_c)

    def prs_b(attivo):
        return Paragraph('<font color="green">✓</font>' if attivo else '<font color="gray">✗</font>', s_cella)

    def pr_bold(v):
        return Paragraph(str(v) if v is not None else '', s_totale)

    def pr_bold_dx(v):
        return Paragraph(f'€ {v:.4f}', s_totale_dx)

    # --- Caricamento dati ---
    contratto_pk = request.GET.get('contratto_pk')
    livello_pk = request.GET.get('livello_pk')
    anno_s = request.GET.get('anno', str(date.today().year))
    try:
        anno = int(anno_s)
    except (ValueError, TypeError):
        anno = date.today().year
    mese_s = request.GET.get('mese', str(date.today().month))
    try:
        mese = int(mese_s)
    except (ValueError, TypeError):
        mese = date.today().month
    codice = request.GET.get('codice', '')

    contratto = None
    if contratto_pk:
        contratto = get_object_or_404(ContrattoAttivo, pk=contratto_pk)

    if livello_pk:
        liv = get_object_or_404(Livello, pk=livello_pk)
        parametri = ParametriCCNL.objects.filter(livello=liv, anno=anno).first()
    elif contratto and hasattr(contratto, 'parametri_minimi') and contratto.parametri_minimi:
        parametri = contratto.parametri_minimi
    else:
        parametri = None

    if not parametri:
        return JsonResponse({'errore': 'Parametri CCNL non trovati'})

    # --- Costruisci voci CCNL ---
    groups, all_toggleable, scatto_value, scatto_effettivo, scatti_attiva, paga_base, paga_base_effettiva, retrib_sost_value, retrib_sost_attiva, bienni_accumulati = _costruisci_voci_ccnl(contratto, parametri)

    # --- Toggles (internal keys per _calcola_busta_data) ---
    toggles_dict = _parse_toggles(request) or {}
    CCNL_TO_INTERNAL = {
        'ind_funzione_conviventi': 'ind_funzione',
        'ind_minori_6_anni_ft': 'ind_bambini_6',
        'ind_assistenza_piu_persone_ft': 'ind_piu_assistiti',
        'ind_certificazione_qualita': 'ind_cert_qualita',
    }
    for ccnl_key, internal_key in CCNL_TO_INTERNAL.items():
        v = request.GET.get(ccnl_key)
        if v is not None:
            toggles_dict[internal_key] = v == '1'

    EXTRA_IND = [
        ('ind_piu_persone_non_conv', 'ind_assistenza_piu_persone_non_conv'),
        ('ind_minori_non_conv', 'ind_minori_6_anni_non_conv'),
        ('ind_piu_persone_qualita', 'ind_piu_persone_qualita'),
        ('ind_minori_qualita', 'ind_minori_qualita'),
        ('ind_assistenza_piu_persone_pt', 'ind_assistenza_piu_persone_pt'),
    ]
    extra_indennita = []
    extra_totale = 0.0
    for toggle_key, param_field in EXTRA_IND:
        v = request.GET.get(toggle_key)
        attiva = (v == '1') if v is not None else (getattr(contratto, toggle_key, False) if contratto else False)
        if attiva:
            val = float(getattr(parametri, param_field, 0) or 0)
            extra_indennita.append({'label': toggle_key, 'importo': val})
            extra_totale += val

    # --- Ore simulate e busta ---
    ore_toggles = {}
    for popup_key, model_key in _ORE_TOGGLE_MAP.items():
        v = request.GET.get(popup_key)
        if v is not None:
            ore_toggles[model_key] = v == '1'
    v_tfr = request.GET.get('rateo_tfr')
    v_ant70 = request.GET.get('rateo_anticipo_70')
    if v_tfr is not None:
        ore_toggles['rateo_tfr'] = v_tfr == '1'
    if v_ant70 is not None:
        ore_toggles['rateo_anticipo_70'] = v_ant70 == '1'
    for cf, _ in _IND_MAPPING_ORE:
        v = request.GET.get(cf)
        if v is not None:
            ore_toggles[cf] = v == '1'

    ore_simulate = _calcola_ore_simulate(contratto, parametri, ore_toggles, scatto_effettivo) if contratto else 0
    busta_data = _calcola_busta_data(contratto, mese, anno, parametri_override=parametri, ore_override=ore_simulate) if contratto else {}
    if busta_data and not busta_data.get('errore') and contratto and ore_simulate:
        ore_reali = float(contratto.ore_mensili_calcolate)
        po = busta_data.get('paga_oraria_lorda', float(parametri.paga_base))
        busta_data['differenza'] = round((ore_simulate - ore_reali) * po, 2)

    if extra_indennita and busta_data:
        busta_data.setdefault('indennita', []).extend(extra_indennita)
        busta_data['totale_indennita'] = round(busta_data.get('totale_indennita', 0) + extra_totale, 4)
        busta_data['totale_lordo'] = round(busta_data.get('totale_lordo', 0) + extra_totale, 2)
        busta_data['netto'] = round(busta_data.get('netto', 0) + extra_totale, 4)

    tfr_mode = 'INCLUSO'
    tfr_val = float(parametri.tfr_orario)
    if ore_toggles.get('rateo_tfr') is True:
        tfr_mode = 'INCLUSO'
    elif ore_toggles.get('rateo_anticipo_70') is True:
        tfr_mode = 'ANTICIPO 70%'
        tfr_val = float(parametri.tfr_orario) * 0.7
    elif ore_toggles.get('rateo_tfr') is False:
        tfr_mode = 'SEPARATO'
        tfr_val = 0
    elif contratto:
        mt = contratto.modalita_tfr
        tfr_mode = {'INCLUSO': 'INCLUSO', 'SEPARATO_70': 'ANTICIPO 70%', 'SEPARATO': 'SEPARATO'}.get(mt, mt)
        if tfr_mode == 'ANTICIPO 70%':
            tfr_val = float(parametri.tfr_orario) * 0.7
        elif tfr_mode == 'SEPARATO':
            tfr_val = 0

    # --- Costruzione PDF ---
    elements = []
    elements.append(Paragraph('RIEPILOGO CCNL', s_titolo))
    lab_livello = f'{parametri.livello.codice}' if parametri.livello else ''
    sottotitolo = f'Anno {anno} — Livello {lab_livello}'
    if contratto:
        sottotitolo += f' — {contratto.lavoratore}'
    if codice:
        sottotitolo += f' — Cod. {codice}'
    elements.append(Paragraph(sottotitolo, s_sottotitolo))
    elements.append(HRFlow(width='100%', thickness=1, color=acciaio))
    elements.append(Spacer(1, 6))

    # ---- Sezione: PARAMETRI CCNL ----
    elements.append(Paragraph('PARAMETRI CCNL', s_sezione))
    header = ['Parametro', 'Valore Orario', 'Attivo']
    td = [header]

    # Paga Base
    td.append(['Paga Base Oraria', prd(paga_base), prs(True)])

    # Ratei
    td.append(['Tredicesima Oraria', prd(float(parametri.tredicesima_oraria)),
               prs(ore_toggles.get('paga_13ma', contratto.paga_13ma if contratto else False))])
    td.append(['Ferie Orarie', prd(float(parametri.ferie_orarie)),
               prs(ore_toggles.get('paga_ferie', contratto.paga_ferie if contratto else False))])
    td.append(['Festivi Orari', prd(float(parametri.festivi_orari)),
               prs(ore_toggles.get('paga_festivi', contratto.paga_festivi if contratto else False))])

    tfr_attivo = tfr_mode in ('INCLUSO', 'ANTICIPO 70%')
    tfr_label = f'TFR Orario ({tfr_mode})' if tfr_mode != 'SEPARATO' else 'TFR Orario (SEPARATO)'
    td.append([tfr_label, prd(tfr_val), prs(tfr_attivo)])

    # Indennità (dal popup)
    for cf, pf in _IND_MAPPING_ORE:
        label_map = {
            'ind_certificazione_qualita': 'Certificazione Qualità',
            'ind_piu_persone_non_conv': 'Assistenza Più Persone (NC)',
            'ind_minori_non_conv': 'Minori <6 Anni (NC)',
            'ind_piu_persone_qualita': 'Più Persone + Qualità',
            'ind_minori_qualita': 'Minori + Qualità',
            'ind_assistenza_piu_persone_ft': 'Più Persone (FT)',
            'ind_assistenza_piu_persone_pt': 'Più Persone (PT)',
            'ind_minori_6_anni_ft': 'Minori <6 Anni (FT)',
            'ind_funzione_conviventi': 'Funzione Conviventi',
        }
        val = float(getattr(parametri, pf, 0) or 0)
        if val > 0:
            attiva = ore_toggles.get(cf, getattr(contratto, cf, False) if contratto else False)
            td.append([label_map.get(cf, cf), prd(val), prs(attiva)])

    # Scatti
    if scatto_effettivo > 0:
        td.append([f'Scatti Anzianità ({bienni_accumulati} bienni)', prd(scatto_effettivo),
                   prs(ore_toggles.get('applica_scatti', contratto.applica_scatti if contratto else False))])

    tbl = Table(td, colWidths=[160, 80, 50], repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), acciaio),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#ffffff')),
        ('GRID', (0, 0), (-1, -1), 0.5, grigio_bordo),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 12))

    # ---- Sezione: RIEPILOGO BUSTA PAGA ----
    if busta_data and not busta_data.get('errore'):
        b = busta_data
        elements.append(Paragraph('RIEPILOGO BUSTA PAGA', s_sezione))
        bh = ['Voce', 'Importo']
        bd = [bh]

        bd.append(['Budget Mensile', pr_bold_dx(b.get('budget_mensile', 0))])
        bd.append(['Ore Mensili', Paragraph(f'{b.get("ore_mensili", 0):.2f}h', s_cella)])
        bd.append(['Paga Base', pr_bold_dx(b.get('paga_base', {}).get('totale', 0) if isinstance(b.get('paga_base'), dict) else 0)])
        bd.append(['Indennità', pr_bold_dx(b.get('totale_indennita', 0))])
        bd.append(['Scatti Anzianità', pr_bold_dx(b.get('scatti_anzianita', {}).get('valore', 0) if isinstance(b.get('scatti_anzianita'), dict) else 0)])
        bd.append(['Ratei Pagati', pr_bold_dx(b.get('totale_ratei_pagati', 0))])
        bd.append(['', ''])
        bd.append(['Totale Lordo', Paragraph(f'€ {b.get("totale_lordo", 0):.2f}', s_sez_busta_dx)])

        contrib = b.get('contributi', {}) or {}
        bd.append(['Contributi INPS', pr_bold_dx(contrib.get('inps', {}).get('totale', 0) if isinstance(contrib.get('inps'), dict) else 0)])
        bd.append(['Contributi Cassa', pr_bold_dx(contrib.get('cassa', {}).get('totale', 0) if isinstance(contrib.get('cassa'), dict) else 0)])
        bd.append(['Totale Contributi', pr_bold_dx(contrib.get('totale', 0))])

        trat = b.get('trattenute', {}) or {}
        bd.append(['Convivenza', pr_bold_dx(trat.get('convivenza', {}).get('totale', 0) if isinstance(trat.get('convivenza'), dict) else 0)])
        bd.append(['Ratei Accantonati', pr_bold_dx(trat.get('ratei_accantonati', 0))])
        bd.append(['', ''])
        bd.append(['NETTO IN BUSTA', Paragraph(f'€ {b.get("netto", 0):.2f}', ParagraphStyle('NettoBusta', parent=s_sez_busta_dx, fontSize=11, textColor=acciaio))])
        bd.append(['TFR Accumulato', pr_bold_dx(b.get('tfr_accumulato', 0))])

        diff = b.get('differenza', 0)
        col_diff = '#34d399' if diff > 0 else ('#f87171' if diff < 0 else '#8A8F98')
        bd.append(['Differenza Budget', Paragraph(f'€ {diff:.2f}', ParagraphStyle('DiffBudget', parent=s_busta_val_dx, fontSize=9, textColor=HexColor(col_diff)))])

        tbl_b = Table(bd, colWidths=[180, 110], repeatRows=1)
        tbl_b.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), acciaio),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#ffffff')),
            ('GRID', (0, 0), (-1, -1), 0.5, grigio_bordo),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW', (0, 7), (-1, 7), 1, acciaio),
            ('LINEBELOW', (0, 13), (-1, 13), 1, acciaio),
        ]))
        elements.append(tbl_b)

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    nome_file = f"Riepilogo_CCNL_{codice}_{anno}.pdf" if codice else "Riepilogo_CCNL.pdf"
    response['Content-Disposition'] = f'inline; filename="{nome_file}"'
    return response
