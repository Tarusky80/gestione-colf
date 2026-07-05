"""Modulo _backup - generato automaticamente da views.py"""

from paghe.views._common_imports import *

logger = logging.getLogger(__name__)


# --- backup_page ---
@login_required
@permesso_richiesto('backup')
@never_cache
def backup_page(request):
    opzioni = get_opzioni()
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        azione = request.POST.get('azione', 'crea')
        try:
            if azione == 'crea':
                tipo = request.POST.get('tipo_backup', 'COMPLETO')
                note = request.POST.get('note_opzionali', '')
                try:
                    backup = GestoreBackup.objects.create(tipo_backup=tipo, note_opzionali=note)
                    if backup.file_json:
                        return JsonResponse({'success': True, 'message': 'Backup creato con successo'})
                    else:
                        return JsonResponse({'success': False, 'message': 'Errore: file backup non generato'})
                except ValueError as e:
                    return JsonResponse({'success': False, 'message': str(e)})
            elif azione == 'ripristina':
                pk = request.POST.get('pk')
                backup = get_object_or_404(GestoreBackup, pk=pk)
                ok = backup.ripristina_comando()
                if ok:
                    return JsonResponse({'success': True, 'message': 'Database ripristinato con successo!'})
                else:
                    return JsonResponse({'success': False, 'message': 'Errore: file backup non trovato.'})
            elif azione == 'elimina':
                pk = request.POST.get('pk')
                backup = get_object_or_404(GestoreBackup, pk=pk)
                for f in (backup.file_json, backup.file_db):
                    if f and os.path.isfile(f):
                        try:
                            os.unlink(f)
                        except Exception:
                            logger.warning("Impossibile eliminare file backup: %s", f)
                backup.delete()
                return JsonResponse({'success': True, 'message': 'Backup eliminato.'})
            elif azione == 'pulisci':
                from datetime import timedelta
                soglia = timezone.now() - timedelta(days=180)
                vecchi = GestoreBackup.objects.filter(data_creazione__lt=soglia)
                conteggio = vecchi.count()
                for b in vecchi:
                    for f in (b.file_json, b.file_db):
                        if f and os.path.isfile(f):
                            try:
                                os.unlink(f)
                            except Exception:
                                logger.exception("Errore in backup_page")
                                pass
                vecchi.delete()
                return JsonResponse({'success': True, 'message': f'{conteggio} backup eliminati (più vecchi di 6 mesi).'})
            elif azione == 'importa-excel':
                import json as json_lib
                import openpyxl
                import tempfile
                from django.core.management import call_command
                file = request.FILES.get('file')
                if not file:
                    return JsonResponse({'success': False, 'message': 'Nessun file caricato.'})
                wb = openpyxl.load_workbook(file)
                fixture = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        continue
                    headers = [str(h) if h is not None else '' for h in rows[0]]
                    for row in rows[1:]:
                        if all(v is None for v in row):
                            continue
                        record = {'model': sheet_name, 'pk': None, 'fields': {}}
                        for i, val in enumerate(row):
                            if i >= len(headers):
                                break
                            h = headers[i]
                            if h == 'pk':
                                record['pk'] = val
                            elif h == 'model':
                                continue
                            else:
                                record['fields'][h] = val
                        fixture.append(record)
                if not fixture:
                    return JsonResponse({'success': False, 'message': 'Nessun dato valido trovato nel file Excel.'})
                temp_path = None
                try:
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
                        json_lib.dump(fixture, f, ensure_ascii=False, indent=2)
                        temp_path = f.name
                    call_command('loaddata', temp_path, format='json')
                    return JsonResponse({'success': True, 'message': f'Importati {len(fixture)} record da Excel.'})
                finally:
                    if temp_path and os.path.exists(temp_path):
                        os.unlink(temp_path)
        except Exception as e:
            logger.exception("Errore in backup_page")
            return JsonResponse({'success': False, 'message': f'Errore: {str(e)}'})

    backups = GestoreBackup.objects.all().order_by('-data_creazione')
    context = {
        'opzioni': opzioni,
        'backups': backups,
        'count': backups.count(),
        'tipi_backup': GestoreBackup.TIPO_BACKUP,
    }
    return render(request, 'paghe/backup.html', context)


# --- backup_auto ---
@login_required
def backup_auto(request):
    try:
        backup = GestoreBackup.objects.create(tipo_backup='COMPLETO', note_opzionali='Backup automatico schedulato')
        if backup.file_json:
            return JsonResponse({'success': True, 'message': 'Backup automatico creato con successo.', 'pk': backup.pk})
        else:
            return JsonResponse({'success': False, 'message': 'Errore: file backup non generato.'})
    except Exception as e:
        logger.exception("Errore in backup_auto")
        return JsonResponse({'success': False, 'message': f'Errore: {str(e)}'})


# --- esporta_backup_excel ---
@login_required
def esporta_backup_excel(request, pk):
    import json as json_lib
    import openpyxl
    import os
    from openpyxl.styles import Font, PatternFill
    backup = get_object_or_404(GestoreBackup, pk=pk)
    if not backup.file_json:
        return HttpResponse('Nessun file JSON associato al backup.', status=404)
    if not os.path.isfile(backup.file_json) or os.path.getsize(backup.file_json) == 0:
        return HttpResponse('Il file backup è vuoto o non trovato. Impossibile esportare.', status=400)
    try:
        with open(backup.file_json, encoding='utf-8') as f:
            data = json_lib.load(f)
    except (FileNotFoundError, json_lib.JSONDecodeError):
        return HttpResponse('File backup non trovato o danneggiato.', status=404)
    if not data:
        return HttpResponse('Il backup non contiene dati. Impossibile esportare.', status=400)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='5E6AD2', end_color='5E6AD2', fill_type='solid')
    model_groups = {}
    for entry in data:
        model = entry['model']
        model_groups.setdefault(model, []).append(entry)
    for model_name, entries in model_groups.items():
        ws = wb.create_sheet(title=model_name[:31])
        all_keys = ['pk']
        seen = set()
        for entry in entries:
            for k in entry.get('fields', {}):
                if k not in seen:
                    seen.add(k)
                    all_keys.append(k)
        for col_idx, key in enumerate(all_keys, 1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, entry in enumerate(entries, 2):
            ws.cell(row=row_idx, column=1, value=entry.get('pk'))
            for col_idx, key in enumerate(all_keys[1:], 2):
                val = entry.get('fields', {}).get(key)
                if val is None:
                    continue
                if not isinstance(val, (str, int, float, bool)):
                    val = str(val)
                ws.cell(row=row_idx, column=col_idx, value=val)
        ws.auto_filter.ref = ws.dimensions
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="backup_{backup.pk}_{backup.data_creazione.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# --- ajax_ripristina_da_upload ---
@login_required
@require_http_methods(['POST'])
def ajax_ripristina_da_upload(request):
    import tempfile, os, json as json_lib
    from django.core.management import call_command
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'success': False, 'error': 'Nessun file caricato.'})
    if not file.name.endswith('.json'):
        return JsonResponse({'success': False, 'error': 'Il file deve essere un JSON (.json).'})
    try:
        raw = file.read()
        json_lib.loads(raw)
    except Exception:
        logger.exception("Errore in ajax_ripristina_da_upload")
        return JsonResponse({'success': False, 'error': 'Il file non è un JSON valido.'})
    tmpdir = tempfile.mkdtemp()
    tmp_path = os.path.join(tmpdir, 'upload_ripristino.json')
    try:
        with open(tmp_path, 'wb') as f:
            f.write(raw)
        call_command('loaddata', tmp_path, format='json', verbosity=0)
        return JsonResponse({'success': True, 'message': 'Database ripristinato con successo dal file caricato.'})
    except Exception as e:
        logger.exception("Errore in ajax_ripristina_da_upload")
        return JsonResponse({'success': False, 'error': f'Errore durante il ripristino: {str(e)}'})
    finally:
        try:
            os.remove(tmp_path)
            os.rmdir(tmpdir)
        except Exception:
            logger.warning("Impossibile pulire file temporanei backup: %s", tmp_path)


# --- ajax_ripristina_db ---
@login_required
@permesso_richiesto('backup')
@require_http_methods(['POST'])
def ajax_ripristina_db(request, pk):
    import zipfile, tempfile, shutil
    from django.db import connection
    backup = get_object_or_404(GestoreBackup, pk=pk)
    if not backup.file_db or not os.path.isfile(backup.file_db):
        return JsonResponse({'success': False, 'message': 'File ZIP database non trovato.'})
    db_path = settings.DATABASES['default']['NAME']
    if not db_path:
        return JsonResponse({'success': False, 'message': 'Percorso database non configurato.'})
    tmp_dir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(backup.file_db, 'r') as zf:
            zf.extract('db.sqlite3', tmp_dir)
        extracted = os.path.join(tmp_dir, 'db.sqlite3')
        if not os.path.isfile(extracted):
            return JsonResponse({'success': False, 'message': 'db.sqlite3 non trovato nello ZIP.'})
        connection.close()
        shutil.copy2(extracted, db_path)
        return JsonResponse({'success': True, 'message': 'Database ripristinato con successo! Riavviare il server per sicurezza.'})
    except Exception as e:
        logger.exception("Errore in ajax_ripristina_db")
        return JsonResponse({'success': False, 'message': f'Errore ripristino DB: {str(e)}'})
    finally:
        try:
            shutil.rmtree(tmp_dir)
        except Exception:
            logger.warning("Impossibile rimuovere directory temporanea: %s", tmp_dir)


# --- scarica_backup_json ---
@login_required
def scarica_backup_json(request, pk):
    backup = get_object_or_404(GestoreBackup, pk=pk)
    if not backup.file_json or not os.path.isfile(backup.file_json):
        raise Http404("File JSON non trovato.")
    with open(backup.file_json, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(backup.file_json)}"'
        return response


# --- scarica_backup_db ---
@login_required
def scarica_backup_db(request, pk):
    backup = get_object_or_404(GestoreBackup, pk=pk)
    if not backup.file_db or not os.path.isfile(backup.file_db):
        raise Http404("File ZIP database non trovato.")
    with open(backup.file_db, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(backup.file_db)}"'
        return response
