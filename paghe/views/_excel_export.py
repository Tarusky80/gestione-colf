"""Esportazione Excel filtrata per buste paga e contratti."""
from paghe.views._common_imports import *
from paghe.views._constants import XLSX_EXPORT_MAP
from paghe.views._helpers import _resolve_valore
from paghe.models import BustaPaga, ContrattoLavoro

EXPORT_CONFIG = {
    'buste-paga': {
        'conf_key': 'buste-paga',
        'base_qs': BustaPaga.objects.select_related(
            'contratto__datore', 'contratto__lavoratore', 'contratto__parametri_minimi__livello'
        ),
        'filters': {
            'tipo': 'tipo_calcolo__exact',
            'stato': 'stato__exact',
            'mese': 'mese__exact',
            'anno': 'anno__exact',
            'datore': 'contratto__datore__nome_cognome__icontains',
            'lavoratore': 'contratto__lavoratore__nome_cognome__icontains',
            'contratto_pk': 'contratto__pk__exact',
        }
    },
    'contratti': {
        'conf_key': 'contratti',
        'base_qs': ContrattoLavoro.objects.select_related(
            'datore', 'lavoratore', 'parametri_minimi__livello'
        ).prefetch_related('progetto__tipo'),
        'filters': {
            'stato': 'stato__exact',
            'datore_id': 'datore__pk__exact',
            'lavoratore_id': 'lavoratore__pk__exact',
        }
    },
}

@login_required
@permesso_richiesto('report')
def ajax_esporta_xlsx_filtrato(request, tipo):
    conf_export = EXPORT_CONFIG.get(tipo)
    if not conf_export:
        return JsonResponse({'success': False, 'message': 'Tipo non valido'})
    conf = XLSX_EXPORT_MAP.get(conf_export['conf_key'])
    if not conf:
        return JsonResponse({'success': False, 'message': 'Configurazione non trovata'})
    import openpyxl
    from openpyxl.styles import Font as OpFont, PatternFill as OpFill
    conf['model']
    fields = conf['fields']
    qs = conf_export['base_qs']
    for param, lookup in conf_export['filters'].items():
        val = request.GET.get(param)
        if val:
            qs = qs.filter(**{lookup: val})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = conf['sheet'][:31]
    header_font = OpFont(bold=True, color='FFFFFF', size=11)
    header_fill = OpFill(start_color='5E6AD2', end_color='5E6AD2', fill_type='solid')
    for col_idx, (label, _) in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, obj in enumerate(qs, 2):
        for col_idx, (label, field_path) in enumerate(fields, 1):
            val = _resolve_valore(obj, field_path)
            if isinstance(val, bool):
                val = 'Si' if val else 'No'
            elif isinstance(val, date):
                val = val.strftime('%d/%m/%Y')
            elif val is None or val == '':
                val = ''
            ws.cell(row=row_idx, column=col_idx, value=val)
    ws.auto_filter.ref = ws.dimensions
    oggi_str = date.today().strftime('%Y%m%d')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{tipo}_{oggi_str}.xlsx"'
    wb.save(response)
    return response
